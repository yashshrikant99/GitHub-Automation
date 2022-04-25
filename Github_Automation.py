import json
import os
import pprint
import subprocess
from subprocess import run
import openpyxl
from pathlib import Path
from openpyxl import Workbook
from dotenv import load_dotenv
load_dotenv(override=True)

pre_proxy_config_non_prod =os.getenv('pre-proxy-config-non-prod')
pre_proxy_config_preprod=os.getenv('pre-proxy-config-preprod')
pre_proxy_config_se_prod=os.getenv('pre-proxy-config-se-prod')
post_proxy_config_non_prod=os.getenv('post-proxy-config-non-prod')
post_proxy_config_preprod=os.getenv('post-proxy-config-preprod')
post_proxy_config_se_prod=os.getenv('post-proxy-config-se-prod')

def dir_config():
    env={
        "1":pre_proxy_config_non_prod+"/resources/edge/env/",
        "2":pre_proxy_config_preprod+"/resources/edge/env/",
        "3":pre_proxy_config_se_prod+"/resources/edge/env/",
        "4":post_proxy_config_non_prod+"/resources/edge/org",
        "5":post_proxy_config_preprod+"/resources/edge/org",
        "6":post_proxy_config_se_prod+"/resources/edge/org"
    }
    instances={
        "1":["dev","qa","explore"],
        "2":["perf","preprod"],
        "3":["sandbox","prod"],
        "4":["apiProducts","developerApps","developers"],
        "5":["apiProducts","developerApps","developers"],
        "6":["apiProducts","developerApps","developers"]
    }
    files=["caches","kvms","targetServers"]
    config={}
    for i in env:
        for j in instances[i]:
            if(i<'4'):
                for k in files:
                    config[i+'.'+j+'.'+k]=(env[i]+j+'/'+k+'.json')
            else:
                config[i+'.'+j]=(env[i]+'/'+j+'.json')
    return config,env

def run_commands(commands,path):
    # print(commands)
    result = subprocess.run(commands, stdout=subprocess.PIPE,stderr=subprocess.STDOUT,shell=True, cwd=path)
    result_process = result.stdout.decode("utf-8")
    return result_process

def fetch_latest(path):
    bashCommand = ["git reset --hard","git switch master","git pull origin master","git fetch"]
    for i in bashCommand:
        # print(run_commands(i,path))
        run_commands(i,path)

def checkout(path,branch):
    checkout= run_commands("git checkout -b "+branch, path)
    if("already exists" in checkout):
        print('branch already exists , switching to existing branch')
        print(run_commands("git checkout "+branch, path))
    #we now need to check if the mode is read and write and take decisions accordingly

def push(path,read_write,branch,commit_message):
    if(read_write=='w'):
        run_commands("code .", path)
        save_changes=input('Do you want to save changes? (y/n): ')
        print()
        if(save_changes=='y'):
            bashCommand = ["git add .","git config core.autocrlf true",f'git commit -m "{commit_message}"',f"git push origin {branch}"]
            
            for i in bashCommand:
                print(i)
                print(run_commands(i,path))

def ask_save_changes(read_write,path):
    # if(read_write=='w'):
    #     run_commands("code .", path)
    #     save_changes=input('Do you want to save changes? (y/n): ')
    #     print()
    #     return save_changes
    return

def kvms(read_write,changes,path):
    with open(path, 'r') as f:
        file = json.load(f)
    for i in changes:#this iterates through each kvm name
        check=0
        for j in changes[i]: #this iterates through each value in the kvm to be added/checked
            print(f' searching for "{j["name"]}" with value "{j["value"]}" in {i}')
            for k in file:
                if(i==k['name']): #this is to match the name of the kvm in the main file
                    check=1
                    check2=0
                    for l in k['entry']: #this is to loop through all the values in the 
                        if(j['name']==l['name']): #this is to match the values if they are same
                            check2=1
                            if(j['value']==l['value']):
                                print('\tkvm entry found with exact same values')
                            else:
                                print(f'\t\tno entry found with value "{j["value"]}"')
                                l['value']=j['value'] #this is to update the value of the kvm
                    if(check2==0): # here we need to add the new kvm entry 
                        print(f'\tadding new kvm entry into {i}')
                        k['entry'].append(j)
        if(check==0): #we need to check only once the loop exits since itll throw message for all kvms which dont match
            print(f' no kvm found with name {i}')
    print()
    if(read_write=='w'):
        with open(path, 'w') as json_file:
            json.dump(file, json_file,indent=2)

def open_workbook():
    wb = openpyxl.load_workbook("input.xlsx")
    sheet = wb.active
    max_row,max_column=sheet.max_row, sheet.max_column
    files_to_push={}
    envs_to_push=[]
    for number in range(2,max_column): #this it to populate only the env for which we need to pull latest updates
        values='B'+str(number)
        if sheet[values].value!=None:
            for j in sheet[values].value.split(','):
                if(j[0] not in envs_to_push):
                    envs_to_push.append(j[0])
    print('fetching latest updates from master please wait.....\n')
    for i in envs_to_push:
        fetch_latest(env[i])
    for row in sheet.iter_rows():
        if(row[0].coordinate.endswith('1')): #this is to skip the headers
            continue
        branch=row[0].value
        files=row[1].value
        changes=row[2].value
        read_write=row[3].value
        commit_message=row[4].value
        if(branch!=None and files!=None and changes!=None and read_write!=None):
            if(read_write=='w' and commit_message==None): #this is to make sure that when write mode is enabled we need to have commit msg
                print('please enter commit message')
                continue
            try:
                files=sorted(files.split(','))
                changes=json.loads(changes)
                for i in files:
                    try:
                        if(i[0] not in files_to_push and read_write=='w'):
                            files_to_push[i[0]]={}
                            files_to_push[i[0]]['commit_message']=commit_message
                            files_to_push[i[0]]['branch']=branch
                            files_to_push[i[0]]['path']=env[i[0]]
                        path=config[i]
                        display_path=path.split("/")[2]+"/"+path.split("/")[-2]+"/"+path.split("/")[-1]
                        print(f'***checking {display_path} for branch {branch}')
                        checkout(env[i[0]], branch)
                        kvms(read_write, changes, path)
                        print()
                        if(files.index(i)+1!=len(files)):
                            if(files[files.index(i)+1].split('.')[0]==i.split('.')[0]):
                                continue
                    except KeyError:
                        print(f'{i} is not a valid combination key')
                        continue
            except Exception as e:
                print(f'Exception {e} occured,please check file {path}')
                continue
    return files_to_push

def final_push(files_to_push):
    for i in files_to_push:
        path=files_to_push[i]['path']
        branch=files_to_push[i]['branch']
        commit_message=files_to_push[i]['commit_message']
        push(env[i[0]],read_write='w',branch=files_to_push[i]['branch'],commit_message=files_to_push[i]['commit_message'])

config,env=dir_config()
files_to_push=open_workbook()
final_push(files_to_push)

